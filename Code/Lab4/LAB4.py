from openpyxl import Workbook
from bs4 import BeautifulSoup
import requests as req
import string as st
import random as ran


class Company:
    def __init__(self, name, course, percentage, transactions):
        self.name = name
        self.course = course
        self.percentage = percentage
        self.transactions = transactions


def get_text():
    str_code = st.ascii_lowercase
    result_str = 'q/?s='
    new_code = ''.join(ran.choice(str_code) for j in range(3))
    return result_str + new_code


def get_person(text, soup):
    ref = soup.find(text=text).parent.find("span")
    if ref.text == "":
        return 0
    if text == "Transakcje":
        return int(ref.text.replace(" ", ""))

    return float(ref.text.replace(" ", ""))


def find_course(code):
    link = "https://stooq.pl/" + code
    result = req.get(link)
    result_txt = result.text
    soup = BeautifulSoup(result_txt, "html.parser")
    try:
        course = get_person("Kurs", soup)
        percentage = get_person("Zmiana", soup)
        transactions = get_person("Transakcje", soup)
        name = code.split('=')[1]
        return Company(name, course, percentage, transactions)

    except AttributeError:
        try:
            link = soup.find(id="f16").parent.find('a', href=True)
            if link.text == 'symbol waloru':
                link = get_text()
            else:
                link = link.get('href')

            return find_course(link)

        except AttributeError:
            return find_course(get_text())


def fill_wsGielda(companies):
    i = 1
    for col in wsGielda.iter_cols(min_row=1, max_col=4, max_row=5):
        for cell in col:
            if cell.col_idx == 1:
                cell.value = companies[i].name
            elif cell.col_idx == 2:
                cell.value = companies[i].course
            elif cell.col_idx == 3:
                cell.value = companies[i].percentage
            elif cell.col_idx == 4:
                cell.value = companies[i].transactions
            i += 1
            if i == 5:
                i = 1


def gielda():
    companies = []
    for i in range(5):
        companies.append(find_course(get_text()))
    fill_wsGielda(companies)


def links():
    url = "https://www.youtube.com/"
    result = req.get(url)
    result_txt = result.text
    soup = BeautifulSoup(result_txt, "html.parser")
    ref_list = soup.find_all('a', href=True)
    link_list = []
    for element in ref_list:
        link = element.get("href")
        if url not in link:
            link = "https://www.youtube.com" + link
        link_list.append(link)
    i = 0
    for row in wsLinki.iter_rows(min_row=1, max_row=len(link_list)):
        for cell in row:
            cell.value = link_list[i]
            i += 1


def filmweb():
    url = "https://www.filmweb.pl/film/To%3A+Rozdzia%C5%82+2-2019-793838#"
    result = req.get(url)
    result_txt = result.text
    soup = BeautifulSoup(result_txt, "html.parser")
    wsFilmweb['A1'] = soup.find(itemprop="director").text.strip()
    wsFilmweb['A2'] = soup.find("span", {"class": "block"}).text.strip()
    wsFilmweb['A3'] = soup.find("div", {"class": "filmRating filmRating--hasPanel"}).attrs.get("data-count")
    wsFilmweb['A4'] = soup.find("div", {"class": "filmRating filmRating--hasPanel"}).attrs.get("data-rate")


wb = Workbook()
wsGielda = wb.active
wsGielda.title = "Giełda"
wsLinki = wb.create_sheet("Linki")
wsFilmweb = wb.create_sheet("Filmweb")
gielda()
links()
filmweb()
wb.save('Kostrzynski-175ICB2.xlsx')
